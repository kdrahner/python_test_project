
#### -*- coding: utf-8 -*-
####
#### Created on Fri Apr 24 15:22:09 2020

####@author: kdrah
####

# test_file_04.py
#
# This code reads data in from csv files and writes to a copy of 
# an existing excel file (excel_target_file.xlsx) in a specific location in the file 
# sheet 1 box of cells C3:G9 and M7:P15

# test_data_file_01.csv
# 
#    0,1,2,3,4
#    5,6,7,8,9
#    9,8,7,6,5
#    4,3,2,1,0
#    0,1,2,3,4
#    5,6,7,8,9
#    9,8,7,6,5
#
# test_data_file_02.csv
# 
#   10,11,12,13
#   25,26,27,28
#   39,38,37,36
#   44,43,42,41
#   50,51,52,53
#   65,66,67,68
#   79,78,77,76
#   89,88,87,86
#   94,93,92,91
#
#
import csv
import openpyxl

test_data_file_01 = open('test_data_file_01.csv')
test_data_file_01_Reader = csv.reader(test_data_file_01)
test_data_file_01_Data = list(test_data_file_01_Reader) 
print(test_data_file_01_Data)
print('\n')


c=0
field={}

with open('test_data_file_01.csv', 'r') as csvFile:
    reader = csv.reader(csvFile)
    for row in reader:
        field[c]=row
        print(field[c])
        c=c+1

columns=len (field[0])
rows=len(field)


print('\n')
print('The number of rows: ' + str(rows))
print('The number of columns: ' + str(columns))
print('\n')

# demo accessing a specific piece of data in row/column od csv file. 
# note the data is referenced from 0,0 NOT 1,1
print(test_data_file_01_Data[6][4])


# for rrr in range(row_start_BOX01, row_end_BOX01+1):
#     for ccc in range(col_start_BOX01, col_end_BOX01+1):
#         print('rrr = ', rrr, ' and ccc =', ccc)
#         print('test_data_file_01_Data[',ccc, '][',rrr,'] = ', test_data_file_01_Data[ccc-col_start_BOX01][rrr-row_start_BOX01])

print('\n')
print('*****************************************')        

for rrr in range(0, rows):
    for ccc in range(0, columns):
        #print('rows = ', rrr, ' columns = ',ccc)
        #print(test_data_file_01_Data[rrr][ccc])
        print('data in csv[',rrr,'][',ccc,'] = ', test_data_file_01_Data[rrr][ccc])

print('*****************************************')    
print('\n')



###################################

wbkName = 'excel_target_file.xlsx'


# Enter starting and ending rows and columns
# row 1: 1 column A: 1 
# START BOX #1 row 3:  3  column C: 3 
# END BOX   #1 row 9:  9  column G: 7 
# START BOX #2 row 7:  7  column M: 13 
# ENDBOX    #2 row 15: 15 column P: 16 

row_start_BOX01 = 3
col_start_BOX01 = 3
row_end_BOX01 = 9
col_end_BOX01 = 7

row_start_BOX02 = 7
col_start_BOX02 = 13
row_end_BOX02 = 15
col_end_BOX02 = 16






###################################

wbk = openpyxl.load_workbook(wbkName)
wks = wbk['Sheet1']

# Loop to put content in first BOX 
for rr in range(row_start_BOX01, row_end_BOX01+1):
    for cc in range(col_start_BOX01, col_end_BOX01+1):
        #wks.cell(row=rr, column=cc).value = (rr-1)+(cc-1)
        wks.cell(row=rr, column=cc).value = test_data_file_01_Data[rr-row_start_BOX01][cc-col_start_BOX01]
        
        
# Write text to a given cell
someValue = 'TEXT WRITTEN'
wks.cell(row=20, column=5).value = someValue

# The following shows how you can write to a cell by specifying the cell
# by using the LETTER and NUMBER format for Column and Rows
wks['C20'].value = 33

wks.title = 'renamed_sheet'



# Write results to a new excel file:
wbk.save('New2.xlsx')



wbk.close        
csvFile.close()




